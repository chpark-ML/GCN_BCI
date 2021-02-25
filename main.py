import graph
import network as nt
import utils as ut
import setting as st
import numpy as np
import matplotlib.pyplot as plt
from matplotlib.ticker import MaxNLocator
from matplotlib.pyplot import figure, show
import tensorflow as tf
import os, time, collections, shutil
from sklearn.metrics import accuracy_score, cohen_kappa_score
from openpyxl import Workbook
import openpyxl
# np.random.seed(1)
"""GPU connection"""
import GPUtil
devices = "%d" % GPUtil.getFirstAvailable(order="memory")[0]
os.environ["CUDA_VISIBLE_DEVICES"] = devices

"""set the exp_dir"""
root_dir = st.root_dir
if not os.path.exists(root_dir):
    os.makedirs(root_dir)

"""num of the experiment"""
start_exp = 1
end_exp = 3
start_sbj = 1
end_sbj = 9

"""Open the excel file"""
wb= Workbook()
ws1 = wb.create_sheet('exp_peak', 0)
ws2 = wb.create_sheet('exp_selected', 1)
# ws1 = wb.active
# ws2 = wb.active
# ws2.title = "exp_selected"

for col in range (start_exp, end_exp+1):
    ws1.cell(row=1, column=col+1, value="exp_"+str(col))
    ws2.cell(row=1, column=col+1, value="exp_"+str(col))
for row in range (start_exp, end_sbj + 1):
    ws1.cell(row=row+1, column=1, value="sbj_"+str(row))
    ws2.cell(row=row + 1, column=1, value="sbj_" + str(row))
ws1.cell(row =end_sbj + 2, column = 1, value="avg")
ws2.cell(row =end_sbj + 2, column = 1, value="avg")

wb.save(st.root_dir + "/results.xlsx")

for n_exp in range(start_exp,end_exp+1):
    """exp_setting"""
    # st.min_freq = n_exp

    """save the peak, selected kappa"""
    exp_dir = root_dir + st.exp_dir + "_" + str(n_exp)

    """Set the params"""
    params = dict()
    params['F'] = [256, 480]  # Number of graph convolutional filters.
    k = 3
    params['K'] = [k, k]  # Polynomial orders.
    params['M'] = [512, st.n_class]  # Output dimensionality of fully connected layers.
    params['regularization'] = 5e-4

    params['f_stride'] = 10000
    params['f_range'] = 37

    """variable for result of each sbject"""
    sum_peak = 0
    sum_selected = 0

    """Loop for sbject"""
    for sbj in range(start_sbj, end_sbj + 1):
        tf.reset_default_graph()

        """save the setting"""
        tmp_dir = exp_dir
        if not os.path.exists(tmp_dir):
            os.makedirs(tmp_dir)
        f_set = open(exp_dir + "/setting.txt", 'w')
        f_set.write("min, max : {0}, {1}\n".format(st.min_freq, st.max_freq))
        f_set.write("f_strdie : {0}\n".format(params['f_stride']))
        f_set.write("f_range : {0}\n".format(params['f_range']))
        f_set.close()

        """Initialize strings for data save and load"""
        file_train = 'train'
        file_val = 'val'
        file_test = 'test'

        """load dataset, psd_welch, and save"""
        # train_data, train_label, val_data, val_label = ut.load_dataset(training=True)
        # test_data, test_label = ut.load_dataset(training=False) #(54432, 22, 512, 1), (54432, )
        # ut.save_processed_data(train_data, train_label, fileName= file_train)
        # if st.n_train != 72:
        #     ut.save_processed_data(val_data, val_label, fileName= file_val)
        # ut.save_processed_data(test_data, test_label, fileName= file_test)

        """Load saved psd dataset"""
        dat_train_psd = np.load(st.data_path +"/psd" +"/dat_" + file_train + "_" + str(sbj) + ".npy")
        lbl_train = np.load(st.data_path+"/psd" + "/lbl_" + file_train + "_" + str(sbj) + ".npy")

        dat_val_psd = np.load(st.data_path +"/psd"+ "/dat_" + file_val + "_" + str(sbj) + ".npy")
        lbl_val = np.load(st.data_path +"/psd"+ "/lbl_" + file_val + "_" + str(sbj) + ".npy")

        dat_test_psd = np.load(st.data_path +"/psd"+ "/dat_" + file_test + "_" + str(sbj) + ".npy")
        lbl_test = np.load(st.data_path+"/psd" + "/lbl_" + file_test + "_" + str(sbj) + ".npy")

        """save the PSD data as a image"""
        # for i in range(260):
        #     print(i)
        #     fig = plt.figure()
        #     plt.imshow(np.asarray(dat_train_psd[189* i + 90, :, :]), vmax= 1.0, vmin=0.0)
        #     plt.pcolor
        #     plt.colorbar()
        #     # plt.show()
        #     fig.savefig('plot/'+str(i) + "_plot.png")
        #     plt.close(fig)

        """Select the frequency interest range"""
        dat_train_psd = dat_train_psd[:, :, st.min_freq: st.max_freq+1]
        if st.n_train != 72:
            dat_val_psd = dat_val_psd[:, :, st.min_freq: st.max_freq+1]
        dat_test_psd = dat_test_psd[:, :, st.min_freq: st.max_freq+1]

        """Graph construction"""
        ## define the coordinate for calculating the distance between nodes
        coordinate_channel = np.array([[0,2], [-2,1], [-1,1],[0,1],[1,1],[2,1],[-3,0],[-2,0],[-1,0],[0,0],[1,0],[2,0],[3,0],[-2,-1],[-1,-1],[0,-1],[1,-1],[2,-1],[-1,-2],[0,-2],[1,-2],[0,-3]])
        dist, idx = graph.distance_scipy_spatial(coordinate_channel, k=st.n_knn)
        A = [graph.adjacency(dist, idx).astype(np.float32)]

        """find the relation using illustration"""
        # adj = graph.AdjFromIllust()
        # A = [sparse.csr_matrix(adj)]

        """PlaceHolder"""
        ph_X = tf.placeholder(dtype=tf.float32, shape=[st.batch_size, dat_train_psd.shape[1], dat_train_psd.shape[2]])
        ph_Y = tf.placeholder(dtype=tf.float32, shape=[st.batch_size])
        ph_X_test = tf.placeholder(dtype=tf.float32, shape=[(st.n_timepoint - st.test_win_size + 1), dat_test_psd.shape[1], dat_test_psd.shape[2]])
        ph_Y_test = tf.placeholder(dtype=tf.float32, shape=[(st.n_timepoint - st.test_win_size + 1)])

        """global step"""
        global_step = tf.Variable(0, trainable=False, name='global_step')

        """training model"""
        params['reuse'] = False
        logit, _ , adj_train, loss= nt.cgcnn(input=ph_X, label=ph_Y, L=A, **params)
        variables = tf.get_collection(tf.GraphKeys.TRAINABLE_VARIABLES, scope="EEG_GCN")

        """test model"""
        params["reuse"] = True
        logit_test, pred, adj_test,loss_test= nt.cgcnn(input=ph_X_test,label=ph_Y_test, L=A, **params)

        """learning rate + optimizer"""
        stepForEveryDecay = st.stepForEveryDecay
        rateForDecay = st.rateForDecay
        momentum = st.momentum
        learning_rate = tf.train.exponential_decay(st.learning_rate, global_step, stepForEveryDecay, rateForDecay, staircase=True)
        update_ops = tf.get_collection(tf.GraphKeys.UPDATE_OPS)
        with tf.control_dependencies(update_ops):
            optimizer = tf.train.MomentumOptimizer(learning_rate=learning_rate, momentum=momentum).minimize(loss=loss, var_list=variables, global_step=global_step)
            # optimizer = tf.train.AdamOptimizer(learning_rate).minimize(loss, var_list=variables, global_step = global_step)

        """Session"""

        sess = tf.Session()
        # saver = tf.train.Saver(tf.global_variables())
        sess.run(tf.global_variables_initializer())

        """log directory"""
        log_dir = exp_dir + '/logs/' + str(sbj)
        if not os.path.exists(log_dir):
            os.makedirs(log_dir)
        writer = tf.summary.FileWriter(log_dir, sess.graph)

        """start train"""
        indices = collections.deque()
        num_steps = int(st.num_epochs * dat_train_psd.shape[0] / st.batch_size)
        peak_test_kappa = 0
        peak_val_kappa = 0
        selected_kappa = 0
        selected_step = 0
        adj_previous = np.zeros(shape=(22,22), dtype=np.float32)
        for step in range(1, num_steps + 1):
            # Be sure to have used all the samples before using one a second time.
            if len(indices) < st.batch_size:
                indices.extend(np.random.permutation(dat_train_psd.shape[0]))
            idx = [indices.popleft() for i in range(st.batch_size)]

            ## Set the batch data
            batch_data, batch_labels = dat_train_psd[idx, :, :], lbl_train[idx]
            if type(batch_data) is not np.ndarray:
                batch_data = batch_data.toarray()  # convert sparse matrices

            """training"""
            feed_dict = {ph_X: batch_data, ph_Y: batch_labels}
            _, loss_, currentLR, adj_ = sess.run([optimizer, loss, learning_rate, adj_train], feed_dict=feed_dict)

            """summary"""
            summary = tf.Summary()
            summary.value.add(tag='train_loss', simple_value = loss_)
            writer.add_summary(summary, step)

            # Periodical evaluation of the model.
            if step % st.eval_frequency == 0 or step == num_steps:
                print("--------- Subject Num : " + str(sbj))
                print("--------- global_step : {:d} / {:d}---------".format(step, num_steps + 1))
                flag = False
                """validation"""
                if st.n_train != 72 :
                    prediction = np.zeros(shape=(st.n_val_trial))
                    ground_truth = np.zeros(shape=(st.n_val_trial))
                    for trials in range(0, st.n_val_trial):
                        test_batch_x = dat_val_psd[trials * st.test_n_rolled:(trials + 1) * st.test_n_rolled, :, :]
                        test_batch_y = lbl_val[trials * st.test_n_rolled:(trials + 1) * st.test_n_rolled]
                        pred_ = sess.run([pred], feed_dict={ph_X_test: test_batch_x})
                        pred_ = np.argmax(np.bincount(np.squeeze(np.asarray(pred_)), minlength=st.n_class), axis=-1)
                        ground_truth[trials] = lbl_val[trials * st.test_n_rolled]
                        prediction[trials] = pred_
                    ##  Accuracy
                    # print("validation Accuracy: %f, Kappa value: %f"
                    #       % (accuracy_score(y_true=ground_truth, y_pred=prediction),
                    #          cohen_kappa_score(y1=ground_truth, y2=prediction)))
                    val_kappa = cohen_kappa_score(y1=ground_truth, y2=prediction)

                    """summary"""
                    summary.value.add(tag='val_kappa', simple_value=val_kappa)
                    writer.add_summary(summary, step)

                    if peak_val_kappa <= val_kappa:
                        peak_val_kappa = val_kappa
                        flag = True

                """test"""
                prediction = np.zeros(shape=(st.n_test_trial))
                ground_truth = np.zeros(shape=(st.n_test_trial))
                for trials in range(0, st.n_test_trial):
                    test_batch_x = dat_test_psd[trials * st.test_n_rolled:(trials + 1) * st.test_n_rolled, :, :]
                    test_batch_y = lbl_test[trials * st.test_n_rolled:(trials + 1) * st.test_n_rolled]
                    pred_, adj__  = sess.run([pred, adj_test], feed_dict={ph_X_test: test_batch_x})
                    pred_ = np.argmax(np.bincount(np.squeeze(np.asarray(pred_)), minlength=st.n_class), axis=-1)
                    ground_truth[trials] = lbl_test[trials * st.test_n_rolled]
                    prediction[trials] = pred_

                """Accuracy"""
                accuracy = accuracy_score(y_true=ground_truth, y_pred=prediction)
                kappa = cohen_kappa_score(y1=ground_truth, y2=prediction)

                """summary"""
                # summary = tf.Summary()
                # summary.value.add(tag='test_loss', simple_value=loss__)
                # writer.add_summary(summary, step)
                summary.value.add(tag='test_kappa', simple_value=kappa)
                writer.add_summary(summary, step)
                if peak_test_kappa < kappa:
                    peak_test_kappa = kappa
                if flag ==True:
                    selected_kappa = kappa
                    selected_step = step
                print("test Accuracy: %f, Kappa value: %f" % (accuracy, kappa))
                print("peak Kappa : {0}".format(peak_test_kappa))
                print("Selected Kappa : {0}".format(selected_kappa))
                print("Selected step : {0}".format(selected_step))

                """save the peak, selected kappa"""
                tmp_dir = exp_dir + '/result/' + str(sbj)
                if not os.path.exists(tmp_dir):
                    os.makedirs(tmp_dir)
                f = open(tmp_dir + "/result.txt", 'w')
                f.write("peak Kappa : {0:10.4f}\n".format(peak_test_kappa))
                f.write("Selected Kappa : {0:10.4f}\n".format(selected_kappa))
                f.write("Selected step : {0:10.1f}\n".format(selected_step))
                f.close()

                if st.img_flag == True :
                    """adj matrix"""
                    adj_img = adj__
                    fig, (ax_adj) = plt.subplots()
                    im_adj = ax_adj.imshow(adj_img, vmin=-1.0, vmax=1.0)
                    # plotting
                    x = range(1,23)
                    y = range(1,23)
                    fig.colorbar(im_adj, orientation='vertical')
                    ax_adj.set_xticks(np.arange(len(x)))
                    ax_adj.set_yticks(np.arange(len(y)))
                    ax_adj.set_xticklabels(x)
                    ax_adj.set_yticklabels(y)
                    plt.grid(True)
                    plt.xlabel('channel')
                    plt.ylabel('channel')
                    '''Rotate the tick labels and set their alignment.'''
                    plt.setp(ax_adj.get_xticklabels(), rotation=90, ha="right", rotation_mode="anchor")
                    ax_adj.set_title("The adjacency matrix")

                    """show the plot, save the plot"""
                    plot_dir = exp_dir + '/plot/' + str(sbj) + '/' + 'adj'
                    if not os.path.exists(plot_dir):
                        os.makedirs(plot_dir)
                    save_plot_dir = plot_dir + '/' + str(step) + '_change_adj.png'
                    plt.savefig(save_plot_dir)
                    plt.close('all')

                    """diff matrix"""
                    diff_img = adj__ - adj_previous
                    fig, (ax_diff) = plt.subplots()
                    im_adj = ax_diff.imshow(diff_img, vmin=-0.04, vmax=0.04)

                    """plotting"""
                    x = range(1, 23)
                    y = range(1, 23)
                    fig.colorbar(im_adj, orientation='vertical')
                    ax_diff.set_xticks(np.arange(len(x)))
                    ax_diff.set_yticks(np.arange(len(y)))
                    ax_diff.set_xticklabels(x)
                    ax_diff.set_yticklabels(y)
                    plt.grid(True)
                    plt.xlabel('channel')
                    plt.ylabel('channel')
                    '''Rotate the tick labels and set their alignment.'''
                    plt.setp(ax_adj.get_xticklabels(), rotation=90, ha="right", rotation_mode="anchor")
                    ax_adj.set_title("Change of the adjacency matrix")

                    """show the plot, save the plot"""
                    plot_dir = exp_dir + '/plot/' + str(sbj) + '/' + 'diff'
                    if not os.path.exists(plot_dir):
                        os.makedirs(plot_dir)
                    save_diff_dir = exp_dir + '/plot/' + str(sbj) + '/' + 'diff' + '/' + str(step) + '_change_adj.png'
                    plt.savefig(save_diff_dir)
                    plt.close('all')

                    """update the previous adjacency matrix"""
                    adj_previous = adj__

        ws1.cell(row=sbj + 1, column=n_exp + 1, value="{0:10.4f}".format(peak_test_kappa))

        sum_peak += peak_test_kappa
        ws2.cell(row=sbj + 1, column=n_exp + 1, value="{0:10.4f}".format(selected_kappa))
        sum_selected += selected_kappa
        wb.save(st.root_dir + "/results.xlsx")
        writer.close()
        sess.close()

    ws1.cell(row=11, column=n_exp + 1, value="{0:10.4f}".format(sum_peak/9))
    ws2.cell(row=11, column=n_exp + 1, value="{0:10.4f}".format(sum_selected / 9))
    wb.save(st.root_dir + "/results.xlsx")
wb.close()